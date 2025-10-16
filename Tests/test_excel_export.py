#!/usr/bin/env python3
"""
Test script for Excel export functionality
Tests Excel file creation, structure, and data integrity
"""

import sys
import os
import shutil
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from modules.config_manager import load_env_config, get_output_config
from modules.excel_utils import export_to_excel
from openpyxl import load_workbook


def create_test_data():
    """Create sample car data for testing"""
    return [
        {
            'Brand': 'Toyota',
            'Model': 'Corolla',
            'Production Date': 'юни 2007',
            'Price_EUR': 2964.98,
            'Price_BGN': 5799,
            'Engine': '124 к.с.',
            'Fuel Type': 'Бензинов',
            'Transmission': 'Ръчна',
            'Mileage': '303000 км',
            'Color': 'Сив',
            'Location': 'София',
            'Phone': '0893911291',
            'Link': 'https://www.mobile.bg/test1',
            'Описание': 'Добре поддържана кола',
            'Car Extras': 'Климатик, ABS'
        },
        {
            'Brand': 'BMW',
            'Model': 'X3',
            'Production Date': 'януари 2015',
            'Price_EUR': 15000.00,
            'Price_BGN': 29300,
            'Engine': '190 к.с.',
            'Fuel Type': 'Дизелов',
            'Transmission': 'Автоматична',
            'Mileage': '150000 км',
            'Color': 'Черен',
            'Location': 'Пловдив',
            'Phone': '0888123456',
            'Link': 'https://www.mobile.bg/test2',
            'Описание': 'Отлично състояние',
            'Car Extras': 'Кожени седалки, Навигация'
        },
        {
            'Brand': 'Audi',
            'Model': 'A4',
            'Production Date': 'май 2018',
            'Price_EUR': 22500.50,
            'Price_BGN': 44000,
            'Engine': '150 к.с.',
            'Fuel Type': 'Бензинов',
            'Transmission': 'Ръчна',
            'Mileage': '95000 км',
            'Color': 'Бял',
            'Location': 'Варна',
            'Phone': '0899987654',
            'Link': 'https://www.mobile.bg/test3',
            'Описание': 'Като нова',
            'Car Extras': 'LED фарове, Алуминиеви джанти'
        }
    ]


def test_excel_auto_creation():
    """Test Excel file auto-creation from .env config"""
    print('=== TESTING EXCEL AUTO-CREATION ===')
    
    # Clean up any existing test files
    if os.path.exists('docs'):
        shutil.rmtree('docs')
        print('🗑️ Removed existing docs directory')
    
    # Load configuration
    load_env_config()
    config = get_output_config()
    
    print(f'\n📋 Configuration loaded:')
    for key, value in config.items():
        print(f'  {key}: {value}')
    
    # Check initial state
    print(f'\n📂 Initial state:')
    print(f'  docs directory exists: {os.path.exists("docs")}')
    print(f'  Excel file exists: {os.path.exists(config["excel_path"])}')
    
    # Create test data and export
    test_cars = create_test_data()
    print(f'\n📊 Creating test data with {len(test_cars)} cars...')
    
    # Test Excel export with .env defaults
    print(f'\n📊 Testing Excel export...')
    result_path = export_to_excel(test_cars)
    
    # Verify results
    print(f'\n✅ Post-export verification:')
    print(f'  Result path: {result_path}')
    print(f'  File exists: {os.path.exists(result_path)}')
    
    if os.path.exists(result_path):
        file_size = os.path.getsize(result_path)
        print(f'  File size: {file_size} bytes')
        return True
    else:
        print('❌ Excel file was not created')
        return False


def test_excel_structure():
    """Test Excel file structure and data integrity"""
    print('\n=== TESTING EXCEL STRUCTURE ===')
    
    try:
        wb = load_workbook('docs/car-data.xlsx')
        print(f'📋 Available sheets: {wb.sheetnames}')
        
        # Use the first sheet
        ws = wb[wb.sheetnames[0]]
        sheet_name = wb.sheetnames[0]
        print(f'\n--- Sheet: {sheet_name} ---')
        print(f'📊 Total rows: {ws.max_row} (including header)')
        print(f'📊 Total columns: {ws.max_column}')
        print(f'📊 Data rows: {ws.max_row - 1} cars')
        
        # Get and verify headers
        headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
        expected_headers = [
            'Brand', 'Model', 'Production Date', 'Price_EUR', 'Price_BGN', 'Engine', 'Fuel Type', 
            'Transmission', 'Mileage', 'Color', 'Location', 'Phone', 
            'Link', 'Описание', 'Car Extras'
        ]
        
        print(f'\n📋 Headers verification:')
        headers_ok = True
        for i, (expected, actual) in enumerate(zip(expected_headers, headers), 1):
            match = expected == actual
            status = '✅' if match else '❌'
            print(f'  {i:2d}. {expected} → {actual} {status}')
            if not match:
                headers_ok = False
        
        # Check price columns data (first 3 cars)
        print(f'\n💰 Price data verification:')
        price_data_ok = True
        
        for row in range(2, min(5, ws.max_row + 1)):
            brand = ws.cell(row=row, column=1).value
            model = ws.cell(row=row, column=2).value
            production_date = ws.cell(row=row, column=3).value  
            price_eur = ws.cell(row=row, column=4).value  
            price_bgn = ws.cell(row=row, column=5).value  
            
            print(f'\n  Car {row-1}: {brand} {model}')
            print(f'    Production Date: {production_date}')
            print(f'    EUR: {price_eur} (type: {type(price_eur).__name__})')
            print(f'    BGN: {price_bgn} (type: {type(price_bgn).__name__})')
            
            # Verify price types
            if not isinstance(price_eur, (int, float)) or not isinstance(price_bgn, (int, float)):
                price_data_ok = False
                print(f'    ❌ Price types incorrect')
            else:
                print(f'    ✅ Price types correct')
        
        wb.close()
        
        return headers_ok and price_data_ok
        
    except Exception as e:
        print(f'❌ Error verifying Excel structure: {e}')
        return False


def test_new_sheet_creation():
    """Test creating new sheets in existing Excel file"""
    print('\n=== TESTING NEW SHEET CREATION ===')
    
    test_cars = create_test_data()[:1]  # Just one car for testing
    
    # Test with custom sheet name
    result_path = export_to_excel(test_cars, sheet_name='TestSheet')
    print(f'✅ Created new sheet, file: {result_path}')
    
    # Verify the sheet exists
    wb = load_workbook(result_path)
    print(f'📋 Sheets after test: {wb.sheetnames}')
    
    success = 'TestSheet' in wb.sheetnames
    wb.close()
    
    if success:
        print('✅ New sheet creation test PASSED')
    else:
        print('❌ New sheet creation test FAILED')
    
    return success


if __name__ == '__main__':
    print('🧪 EXCEL EXPORT TEST SUITE')
    print('=' * 50)
    
    success1 = test_excel_auto_creation()
    success2 = test_excel_structure()
    success3 = test_new_sheet_creation()
    
    print('\n' + '=' * 50)
    if success1 and success2 and success3:
        print('🎉 All Excel export tests PASSED!')
    else:
        print('❌ Some Excel tests FAILED')
        sys.exit(1)