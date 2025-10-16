#!/usr/bin/env python3
"""
Test script for Excel structure verification
Tests Excel file structure, headers, and data integrity
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook


def test_excel_structure():
    """Test Excel structure and data integrity with the latest export"""
    print('ğŸ“Š EXCEL STRUCTURE VERIFICATION TEST')
    print('=' * 50)
    
    excel_path = 'docs/car-data.xlsx'
    if not os.path.exists(excel_path):
        print('âŒ Excel file not found!')
        return False
    
    try:
        print(f'ğŸ“ Opening: {excel_path}')
        
        wb = load_workbook(excel_path)
        print(f'ğŸ“‹ Available sheets: {wb.sheetnames}')
        
        # Test the first sheet (should be the latest)
        ws = wb[wb.sheetnames[0]]
        sheet_name = wb.sheetnames[0]
        
        print(f'\n--- Sheet: {sheet_name} ---')
        print(f'ğŸ“Š Total rows: {ws.max_row} (including header)')
        print(f'ğŸ“Š Total columns: {ws.max_column}')
        print(f'ğŸ“Š Data rows: {ws.max_row - 1} cars')
        
        # Verify headers
        headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
        expected_headers = [
            'Brand', 'Model', 'Production Date', 'Price_EUR', 'Price_BGN', 
            'Engine', 'Fuel Type', 'Transmission', 'Mileage', 'Color', 
            'Location', 'Phone', 'Link', 'ĞĞ¿Ğ¸ÑĞ°Ğ½Ğ¸Ğµ', 'Car Extras'
        ]
        
        print(f'\nğŸ“‹ HEADER VERIFICATION:')
        headers_correct = True
        for i, (expected, actual) in enumerate(zip(expected_headers, headers), 1):
            match = expected == actual
            status = 'âœ…' if match else 'âŒ'
            print(f'  {i:2d}. {expected:<15} â†’ {actual:<15} {status}')
            if not match:
                headers_correct = False
        
        # Verify we have exactly 15 columns (optimized structure)
        if ws.max_column == 15:
            print(f'\nâœ… Column count correct: {ws.max_column} columns')
        else:
            print(f'\nâŒ Column count incorrect: Expected 15, got {ws.max_column}')
            headers_correct = False
        
        # Sample data verification
        print(f'\nğŸ“Š SAMPLE DATA VERIFICATION:')
        data_correct = True
        sample_size = min(5, ws.max_row - 1)
        
        for row in range(2, 2 + sample_size):
            brand = ws.cell(row=row, column=1).value
            model = ws.cell(row=row, column=2).value
            prod_date = ws.cell(row=row, column=3).value
            price_eur = ws.cell(row=row, column=4).value
            price_bgn = ws.cell(row=row, column=5).value
            location = ws.cell(row=row, column=11).value
            
            print(f'\n  Car {row-1}: {brand} {model}')
            print(f'    ğŸ“… Production Date: "{prod_date}"')
            print(f'    ğŸ’° EUR: {price_eur} ({type(price_eur).__name__})')
            print(f'    ğŸ’° BGN: {price_bgn} ({type(price_bgn).__name__})')
            print(f'    ğŸ“ Location: {location}')
            
            # Verify data types
            if isinstance(price_eur, (int, float)) and isinstance(price_bgn, (int, float)):
                print('    âœ… Price data types correct')
            else:
                print('    âŒ Price data types incorrect')
                data_correct = False
            
            # Check production date format
            if prod_date and isinstance(prod_date, str):
                bulgarian_months = ['ÑĞ½ÑƒĞ°Ñ€Ğ¸', 'Ñ„ĞµĞ²Ñ€ÑƒĞ°Ñ€Ğ¸', 'Ğ¼Ğ°Ñ€Ñ‚', 'Ğ°Ğ¿Ñ€Ğ¸Ğ»', 'Ğ¼Ğ°Ğ¹', 'ÑĞ½Ğ¸', 
                                  'ÑĞ»Ğ¸', 'Ğ°Ğ²Ğ³ÑƒÑÑ‚', 'ÑĞµĞ¿Ñ‚ĞµĞ¼Ğ²Ñ€Ğ¸', 'Ğ¾ĞºÑ‚Ğ¾Ğ¼Ğ²Ñ€Ğ¸', 'Ğ½Ğ¾ĞµĞ¼Ğ²Ñ€Ğ¸', 'Ğ´ĞµĞºĞµĞ¼Ğ²Ñ€Ğ¸']
                if any(month in prod_date.lower() for month in bulgarian_months):
                    print('    âœ… Production date format correct')
                else:
                    print(f'    âš ï¸  Production date format may be incomplete: "{prod_date}"')
            else:
                print('    âŒ Production date missing or invalid')
                data_correct = False
        
        wb.close()
        
        # Final result
        overall_success = headers_correct and data_correct
        
        print(f'\nğŸ“Š VERIFICATION RESULTS:')
        print(f'  ğŸ“‹ Headers: {"âœ… PASS" if headers_correct else "âŒ FAIL"}')
        print(f'  ğŸ“Š Data: {"âœ… PASS" if data_correct else "âŒ FAIL"}')
        
        if overall_success:
            print('\nğŸ‰ Excel structure verification PASSED!')
        else:
            print('\nâŒ Excel structure verification FAILED!')
        
        return overall_success
        
    except Exception as e:
        print(f'âŒ Error during Excel verification: {e}')
        return False


if __name__ == '__main__':
    success = test_excel_structure()
    sys.exit(0 if success else 1)