#!/usr/bin/env python3
"""
Test script for the complete crawler system
Tests end-to-end functionality including crawling, extraction, and Excel export
"""

import sys
import os
import shutil
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from modules.config_manager import load_env_config, get_output_config
from modules.logger_config import setup_logging
from modules.url_builder import build_mobilebg_search_url
from modules.web_scraper import get_all_listing_links
from modules.extractors import extract_car_info_mobile
from modules.excel_utils import export_to_excel
from openpyxl import load_workbook


def test_complete_crawler_flow():
    """Test the complete crawler workflow"""
    print('=== TESTING COMPLETE CRAWLER FLOW ===')
    
    # Clean up previous test files
    if os.path.exists('docs'):
        shutil.rmtree('docs')
        print('🗑️ Cleaned up previous test files')
    
    # 1. Load configuration
    print('\n📋 Step 1: Loading configuration...')
    load_env_config()
    config = get_output_config()
    logger = setup_logging()
    print('✅ Configuration loaded')
    
    # 2. Build search URL
    print('\n🔗 Step 2: Building search URL...')
    search_url = build_mobilebg_search_url()
    print(f'✅ Search URL: {search_url}')
    
    # 3. Get car listing links (limited for testing)
    print('\n🕷️ Step 3: Getting car listing links...')
    links = get_all_listing_links(search_url, delay=0.5, max_pages=1, logger=logger)
    print(f'✅ Found {len(links)} car links')
    
    if len(links) == 0:
        print('❌ No car links found - test FAILED')
        return False
    
    # 4. Extract data from first few cars
    print('\n🚗 Step 4: Extracting car data...')
    cars_data = []
    test_links = list(links)[:3]  # Test with first 3 cars
    
    for i, url in enumerate(test_links, 1):
        print(f'  Extracting car {i}/{len(test_links)}...')
        try:
            car_info = extract_car_info_mobile(url, timeout=10)
            if car_info and car_info.get('Brand'):
                cars_data.append(car_info)
                print(f'    ✅ {car_info.get("Brand")} {car_info.get("Model")} - EUR: {car_info.get("Price_EUR")} BGN: {car_info.get("Price_BGN")}')
            else:
                print(f'    ⚠️ No data extracted from {url}')
        except Exception as e:
            print(f'    ❌ Error extracting from {url}: {e}')
    
    if len(cars_data) == 0:
        print('❌ No car data extracted - test FAILED')
        return False
    
    print(f'✅ Extracted data from {len(cars_data)} cars')
    
    # 5. Export to Excel
    print('\n📊 Step 5: Exporting to Excel...')
    result_path = export_to_excel(cars_data)
    print(f'✅ Exported to: {result_path}')
    
    # 6. Verify Excel file
    print('\n🔍 Step 6: Verifying Excel file...')
    if not os.path.exists(result_path):
        print('❌ Excel file not found - test FAILED')
        return False
    
    wb = load_workbook(result_path)
    ws = wb[wb.sheetnames[0]]
    
    print(f'  📊 Excel rows: {ws.max_row} (including header)')
    print(f'  📊 Excel columns: {ws.max_column}')
    print(f'  📊 Data rows: {ws.max_row - 1}')
    
    # Check if we have the expected columns
    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    required_headers = ['Brand', 'Model', 'Production Date', 'Price_EUR', 'Price_BGN']
    headers_ok = all(header in headers for header in required_headers)
    
    if headers_ok:
        print('✅ Required headers found')
    else:
        print('❌ Missing required headers')
        wb.close()
        return False
    
    # Check if price data is properly formatted
    price_data_ok = True
    if ws.max_row > 1:  # Has data
        for row in range(2, min(4, ws.max_row + 1)):  # Check first few rows
            price_eur = ws.cell(row=row, column=4).value  # Price_EUR (column 4)
            price_bgn = ws.cell(row=row, column=5).value  # Price_BGN (column 5)
            
            if not isinstance(price_eur, (int, float)) or not isinstance(price_bgn, (int, float)):
                price_data_ok = False
                break
    
    wb.close()
    
    if price_data_ok:
        print('✅ Price data properly formatted')
    else:
        print('❌ Price data not properly formatted')
        return False
    
    print('\n🎉 Complete crawler flow test PASSED!')
    return True


def test_error_handling():
    """Test error handling in crawler components"""
    print('\n=== TESTING ERROR HANDLING ===')
    
    # Test with invalid URL
    print('🧪 Testing invalid URL handling...')
    try:
        invalid_data = extract_car_info_mobile('https://www.mobile.bg/invalid-url')
        if not invalid_data or invalid_data == {}:
            print('✅ Invalid URL handled correctly (returned empty data)')
        else:
            print('❌ Invalid URL should return empty data')
            return False
    except Exception as e:
        print(f'✅ Invalid URL handled with exception: {e}')
    
    # Test Excel export with empty data
    print('🧪 Testing empty data export...')
    try:
        empty_result = export_to_excel([])
        if empty_result:
            print('✅ Empty data export handled (created file with headers only)')
        else:
            print('❌ Empty data export failed')
            return False
    except Exception as e:
        print(f'❌ Empty data export raised exception: {e}')
        return False
    
    print('✅ Error handling tests PASSED')
    return True


def performance_summary():
    """Display performance summary of the crawler"""
    print('\n=== PERFORMANCE SUMMARY ===')
    
    try:
        # Load the most recent Excel file
        if os.path.exists('docs/car-data.xlsx'):
            wb = load_workbook('docs/car-data.xlsx')
            ws = wb[wb.sheetnames[0]]
            total_cars = ws.max_row - 1
            wb.close()
            
            print(f'📊 Latest crawl results:')
            print(f'  🚗 Total cars processed: {total_cars}')
            print(f'  📁 Excel file: docs/car-data.xlsx')
            print(f'  📋 Sheet: {wb.sheetnames[0]}')
            print(f'  📈 Columns: {ws.max_column} (including Price_EUR and Price_BGN)')
        else:
            print('📊 No recent crawl data available')
            
    except Exception as e:
        print(f'⚠️ Could not analyze performance: {e}')


if __name__ == '__main__':
    print('🧪 COMPLETE CRAWLER TEST SUITE')
    print('=' * 60)
    
    success1 = test_complete_crawler_flow()
    success2 = test_error_handling()
    
    performance_summary()
    
    print('\n' + '=' * 60)
    if success1 and success2:
        print('🎉 ALL COMPLETE CRAWLER TESTS PASSED!')
        print('🚀 System is ready for production use!')
    else:
        print('❌ Some complete crawler tests FAILED')
        sys.exit(1)