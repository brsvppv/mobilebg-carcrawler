"""
Excel utilities for AutoGetCars Crawler
Handles Excel export and formatting functionality
"""

import os
import sys
import logging
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from modules.excel_table_utils import expand_table_to_fit


def write_rows_to_excel(excel_path, sheet_name, data, headers, key_map):
    """
    Write car data rows to Excel file with proper formatting.
    
    Args:
        excel_path (str): Path to Excel file
        sheet_name (str): Name of the worksheet
        data (list): List of car data dictionaries
        headers (list): List of column headers
        key_map (dict): Mapping of headers to data keys
    """
    # Create directory if it doesn't exist
    excel_dir = os.path.dirname(excel_path)
    if excel_dir and not os.path.exists(excel_dir):
        os.makedirs(excel_dir, exist_ok=True)
    
    # Create workbook/sheet if needed
    if not os.path.exists(excel_path):
        # Create new workbook with headers
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name if sheet_name else "Cars"
        ws.append(headers)
        wb.save(excel_path)
        print(f"📝 Created new Excel file: {excel_path}")
    
    try:
        wb = load_workbook(excel_path)
    except Exception as e:
        logging.error(f"Error loading Excel file {excel_path}: {e}")
        # Fallback: create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name if sheet_name else "Cars"
        ws.append(headers)
        wb.save(excel_path)
        print(f"📝 Recreated Excel file: {excel_path}")
    
    wb = load_workbook(excel_path)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        # Create new sheet if needed
        if sheet_name and sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(headers)  # Add headers to new sheet
            print(f"📋 Created new sheet: {sheet_name}")
        else:
            ws = wb.active
    
    # Clear old data except header
    ws.delete_rows(2, ws.max_row)
    
    # Write new data
    for row in data:
        ws.append([row.get(key_map.get(h, h), "") for h in headers])
    
    # Ensure a table exists and covers all data
    if ws.max_row > 1:  # Only if there's data
        try:
            table_name = f"Table_{sheet_name}" if sheet_name else "Table1"
            # Remove existing tables to avoid conflicts
            table_names = list(ws.tables.keys())
            for table_name_to_remove in table_names:
                del ws.tables[table_name_to_remove]
            # Create new table
            table_range = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
            table = Table(displayName=table_name, ref=table_range)
            style = TableStyleInfo(
                name="TableStyleMedium2", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=False
            )
            table.tableStyleInfo = style
            ws.add_table(table)
            # Expand table to fit content
            expand_table_to_fit(ws, table_name=table_name)
        except Exception as e:
            logging.warning(f"Could not create table: {e}")
    
    # Auto-size all columns for better readability
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            if cell.value:
                # Calculate the length needed for this cell
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        
        # Set column width with some padding, but cap it at reasonable size
        adjusted_width = min(max(max_length + 3, 10), 60)  # Min 10, Max 60 chars
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(excel_path)
    return excel_path


def export_to_excel(cars_data, excel_path=None, sheet_name=None):
    """
    Export car data to Excel with proper formatting.
    Uses .env configuration if parameters not provided.
    
    Args:
        cars_data (list): List of car data dictionaries
        excel_path (str, optional): Path to Excel file (uses .env if not provided)
        sheet_name (str, optional): Name of the Excel sheet (uses .env if not provided)
    """
    # Import config manager to get .env settings
    from modules.config_manager import get_output_config
    
    # Get configuration from .env if not provided
    config = get_output_config()
    
    if excel_path is None:
        excel_path = config.get('excel_path', 'docs/car-data.xlsx')
    
    if sheet_name is None:
        sheet_name = config.get('sheet_name', 'CarsData')
    
    # Define standard headers for car data with separate price columns only
    headers = [
        'Brand', 'Model', 'Production Date', 'Price_EUR', 'Price_BGN', 'Engine', 'Fuel Type', 
        'Transmission', 'Mileage', 'Color', 'Location', 'Phone', 
        'Link', 'Описание', 'Car Extras'
    ]
    
    # Create key mapping
    key_map = {h: h for h in headers}
    
    # Ensure the Excel directory exists
    excel_dir = os.path.dirname(excel_path)
    if excel_dir and not os.path.exists(excel_dir):
        os.makedirs(excel_dir, exist_ok=True)
        print(f"📁 Created directory: {excel_dir}")
    
    # Use existing function to write data
    result_path = write_rows_to_excel(
        excel_path=excel_path,
        sheet_name=sheet_name,
        data=cars_data,
        headers=headers,
        key_map=key_map
    )
    
    print(f"📊 Exported {len(cars_data)} cars to: {result_path}")
    return result_path