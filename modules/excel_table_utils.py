from openpyxl.worksheet.table import Table

def expand_table_to_fit(ws, table_name=None):
    """
    Expands the first table (or named table) in the worksheet to cover all non-empty rows and columns.
    """
    if not ws.tables:
        return
    if table_name and table_name in ws.tables:
        table = ws.tables[table_name]
    else:
        table = next(iter(ws.tables.values()))
    if not table:
        return
    min_col = 1
    max_col = ws.max_column
    min_row = 1
    max_row = ws.max_row
    from openpyxl.utils import get_column_letter
    end_col = get_column_letter(max_col)
    table.ref = f"A1:{end_col}{max_row}"
